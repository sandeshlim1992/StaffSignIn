import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import date
from config_manager import load_path, load_password


def generate_staff_sign_in_form(target_date: date):
    """
    Generates a new XLSX file for a specific date in the configured directory.
    """
    workbook_password = load_password()
    save_dir = load_path()
    os.makedirs(save_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Sign In"

    # --- Define Styles ---
    dark_green_fill = PatternFill(start_color="336633", end_color="336633", fill_type="solid")
    white_bold_font_row1 = Font(bold=True, color="FFFFFF", size=24)
    black_bold_font_row2 = Font(bold=True, color="000000", size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # --- Row 1: Main Title and Date ---
    ws.merge_cells('A1:D1')
    ws['A1'] = "STAFF SIGN IN FORM"
    ws['A1'].font = white_bold_font_row1
    ws['A1'].fill = dark_green_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thin_border

    date_str = target_date.strftime("%#m/%#d/%Y")
    ws.merge_cells('E1:E1')
    ws['E1'] = date_str
    ws['E1'].font = white_bold_font_row1
    ws['E1'].fill = dark_green_fill
    ws['E1'].alignment = Alignment(horizontal='right', vertical='center')
    ws['E1'].border = thin_border
    ws.row_dimensions[1].height = 35

    # --- Row 2: Column Headers ---
    headers = ["STAFF NAME", "Clock In", "Clock Out", "Remarks", "All Taps"]
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header_text)
        cell.font = black_bold_font_row2
        cell.border = thin_border
        # MODIFICATION: Changed vertical alignment to 'top'
        cell.alignment = Alignment(horizontal='left', vertical='top')
    ws.row_dimensions[2].height = 25

    # --- Apply Borders and Alignment to data rows ---
    for row_idx in range(3, 203):
        for col_idx in range(1, 6):  # Columns A, B, C, D, E
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            # MODIFICATION: Set default alignment for the row
            alignment_setting = Alignment(vertical='top')

            # MODIFICATION: Apply special wrap text alignment for Column E
            if col_idx == 5:
                alignment_setting = Alignment(vertical='top', wrap_text=True)

            cell.alignment = alignment_setting

    # --- Set Column Widths ---
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30  # Remarks column
    ws.column_dimensions['E'].width = 50

    ws.freeze_panes = 'A3'

    ws.protection.sheet = True
    ws.protection.password = workbook_password

    date_for_filename = target_date.strftime("%#m-%#d-%Y")
    file_path = os.path.join(save_dir, f"{date_for_filename}.xlsx")

    try:
        wb.save(file_path)
        print(f"Successfully generated '{file_path}'.")
        return file_path
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")
        return None