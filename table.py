import os
from PySide6.QtCore import Qt, QTimer, QTime
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QAbstractItemView, QMenu
from PySide6.QtGui import QFont, QColor, QAction
from datetime import datetime

import constants as c
from history_dialog import StaffHistoryDialog
from database_manager import get_taps_for_staff_and_date, log_tap_event
from time_selector_dialog import TimeSelectorDialog

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SignInTable(QTableWidget):
    """
    A custom QTableWidget specialized for displaying and managing the sign-in sheet.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.original_stylesheet = f"""
            QTableWidget {{
                border: none;
                gridline-color: {c.WIN_COLOR_BORDER_LIGHT};
                background-color: {c.WIN_COLOR_WIDGET_BG};
            }}
            QTableWidget::item {{
                border-bottom: 1px solid {c.WIN_COLOR_BORDER_LIGHT};
            }}
            QTableWidget::item:selected {{
                background-color: {c.WIN_COLOR_CONTROL_BG_HOVER};
                color: {c.WIN_COLOR_TEXT_PRIMARY};
            }}
            QHeaderView::section:horizontal {{
                background-color: {c.TABLE_HEADER_BG};
                padding: 10px;
                border: none;
                border-bottom: 1px solid {c.WIN_COLOR_BORDER_LIGHT};
                font-size: 10pt;
                font-weight: bold;
                color: {c.WIN_COLOR_TEXT_PRIMARY};
                min-height: 40px;
            }}
            QTableCornerButton::section {{
                background-color: {c.TABLE_HEADER_BG};
            }}
        """
        self.setStyleSheet(self.original_stylesheet)
        self.verticalHeader().setVisible(False)
        self.verticalHeader().setDefaultSectionSize(36)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)

        self.current_excel_file_path = None

    def display_excel_content(self, file_path):
        if not openpyxl:
            QMessageBox.critical(self, "Missing Dependency", "The 'openpyxl' library is required.")
            return None
        self.current_excel_file_path = file_path
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[2]]

            file_date_str = None
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if isinstance(cell_value, str) and '/' in cell_value:
                    file_date_str = cell_value
                    break
            if not file_date_str:
                file_date_str = sheet['D1'].value if sheet.max_column >= 4 else sheet['C1'].value

            self.setColumnCount(len(headers))
            self.setHorizontalHeaderLabels(headers)
            all_rows = list(sheet.iter_rows(min_row=3, values_only=True))
            self.setRowCount(len(all_rows))
            for row_idx, row_data in enumerate(all_rows):
                for col_idx, cell_value in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                    item.setForeground(QColor(c.WIN_COLOR_TEXT_PRIMARY))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.setItem(row_idx, col_idx, item)

            if self.columnCount() > 3:
                self.setColumnHidden(3, True)
            if self.columnCount() > 4:
                self.setColumnHidden(4, True)

            self.setColumnWidth(0, 350)
            return file_date_str
        except Exception as e:
            QMessageBox.critical(self, "Error Reading File", f"Could not read the Excel file:\\n\\n{e}")
            return None

    def record_swipe(self, file_path, staff_name, current_time_str, total_taps, first_tap_time_str=None):
        if not file_path or not openpyxl:
            return "Error: File path or Excel library not available.", -1

        password = "lsst1234"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            sheet.protection.password = password
            sheet.protection.sheet = False

            target_row_excel = -1
            for row in range(3, sheet.max_row + 200):
                if sheet[f'A{row}'].value == staff_name:
                    target_row_excel = row
                    break

            status_message = ""
            row_to_update = -1

            if target_row_excel == -1:
                next_row_excel = 3
                while sheet[f'A{next_row_excel}'].value is not None:
                    next_row_excel += 1
                row_to_update = next_row_excel
                sheet[f'A{row_to_update}'] = staff_name

                clock_in_time = first_tap_time_str if first_tap_time_str else current_time_str
                sheet[f'B{row_to_update}'] = clock_in_time

                if total_taps % 2 == 0:
                    sheet[f'C{row_to_update}'] = current_time_str
                    status_message = f"Clocked Out: {staff_name}"
                else:
                    status_message = f"Clocked In: {staff_name}"
            else:
                row_to_update = target_row_excel
                if total_taps % 2 == 0:
                    sheet[f'C{row_to_update}'] = current_time_str
                    status_message = f"Clocked Out: {staff_name}"
                else:
                    status_message = f"Tap Recorded: {staff_name}"

            target_row_widget = row_to_update - 3

            existing_taps = sheet[f'E{row_to_update}'].value
            new_taps = f"{existing_taps}, {current_time_str}" if existing_taps else current_time_str
            sheet[f'E{row_to_update}'] = new_taps

            sheet.protection.sheet = True
            workbook.save(file_path)
            return status_message, target_row_widget

        except PermissionError:
            QMessageBox.critical(self, "Permission Denied", "Could not save to Excel file.")
            return "Save failed: Permission Denied.", -1
        except Exception as e:
            QMessageBox.critical(self, "Error Saving File", f"An error occurred while saving to Excel:\\n\\n{e}")
            return "Save failed: An unexpected error occurred.", -1

    def record_manual_entry(self, file_path, staff_name, selected_time, action_type):
        if not file_path or not openpyxl:
            return "Error: File path or Excel library not available.", -1
        password = "lsst1234"
        time_str = selected_time.toString("hh:mm:ss AP")
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            sheet.protection.password = password
            sheet.protection.sheet = False
            target_row_excel, is_clocked_out = -1, False
            for row in range(3, sheet.max_row + 200):
                if sheet[f'A{row}'].value == staff_name:
                    target_row_excel = row
                    if sheet[f'C{row}'].value:
                        is_clocked_out = True
                    break
            status_message, target_row_widget = "", -1
            row_to_update_taps = -1

            if action_type == 'in':
                if target_row_excel != -1:
                    return f"Error: {staff_name} is already on today's sheet.", -1
                next_row_excel = 3
                while sheet[f'A{next_row_excel}'].value is not None:
                    next_row_excel += 1
                sheet[f'A{next_row_excel}'] = staff_name
                sheet[f'B{next_row_excel}'] = time_str
                sheet[f'D{next_row_excel}'] = "Manually Clocked In"
                status_message, target_row_widget = f"Clocked In: {staff_name}", next_row_excel - 3
                row_to_update_taps = next_row_excel

            elif action_type == 'out':
                if target_row_excel == -1:
                    return f"Error: {staff_name} has not clocked in today.", -1
                if is_clocked_out:
                    return f"Error: {staff_name} is already clocked out.", -1
                sheet[f'C{target_row_excel}'] = time_str

                existing_remark = sheet[f'D{target_row_excel}'].value
                if existing_remark == "Manually Clocked In":
                    sheet[f'D{target_row_excel}'] = "Manually Clocked In/Out"
                else:
                    sheet[f'D{target_row_excel}'] = "Manually Clocked Out"

                status_message, target_row_widget = f"Clocked Out: {staff_name}", target_row_excel - 3
                row_to_update_taps = target_row_excel

            if row_to_update_taps != -1:
                existing_taps = sheet[f'E{row_to_update_taps}'].value
                new_taps = f"{existing_taps}, {time_str}" if existing_taps else time_str
                sheet[f'E{row_to_update_taps}'] = new_taps

            sheet.protection.sheet = True
            workbook.save(file_path)
            return status_message, target_row_widget
        except Exception as e:
            QMessageBox.critical(self, "Error Saving File", f"An error occurred while saving to Excel:\\n\\n{e}")
            return f"Save failed: {e}", -1

    # MODIFICATION: Removed the faulty logic that created new blank items.
    def highlight_row(self, row_index, color):
        if row_index < 0 or row_index >= self.rowCount(): return
        for col in range(self.columnCount()):
            item = self.item(row_index, col)
            # Only color the item if it exists.
            if item:
                item.setBackground(color)
                item.setForeground(QColor(c.WIN_COLOR_ACCENT_TEXT_ON_PRIMARY))
        QTimer.singleShot(1200, lambda: self.end_highlight(row_index))

    def end_highlight(self, row_index):
        default_bg_color = QColor(c.WIN_COLOR_WIDGET_BG)
        default_fg_color = QColor(c.WIN_COLOR_TEXT_PRIMARY)
        for col in range(self.columnCount()):
            item = self.item(row_index, col)
            if item:
                item.setBackground(default_bg_color)
                item.setForeground(default_fg_color)

    def get_staff_in_building(self):
        staff_in = []
        for row in range(self.rowCount()):
            name_item = self.item(row, 0)
            if not name_item or not name_item.text():
                continue

            cin_item = self.item(row, 1)
            cout_item = self.item(row, 2)

            if cin_item and cin_item.text() and (not cout_item or not cout_item.text()):
                staff_in.append(name_item.text())
        return staff_in

    def clear_table(self):
        self.setRowCount(0)
        self.clearContents()

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        clock_in_action = QAction("Manually Clock-In/Out", self)
        clock_in_action.triggered.connect(self._manually_clock_in_out)
        menu.addAction(clock_in_action)
        if self.selectedIndexes():
            menu.addSeparator()
            delete_action = QAction("Delete Entry", self)
            delete_action.triggered.connect(self._delete_selected_entry)
            menu.addAction(delete_action)
            view_history_action = QAction("View History", self)
            view_history_action.triggered.connect(self._view_history_selected_entry)
            menu.addAction(view_history_action)
        menu.exec(event.globalPos())

    def _manually_clock_in_out(self):
        dialog = TimeSelectorDialog(self)
        if dialog.exec():
            staff, time, action = dialog.get_selected_staff(), dialog.get_selected_time(), dialog.action
            if not staff:
                QMessageBox.warning(self, "Selection Error", "No staff member was selected.")
                return
            staff_name, token = staff['name'], staff['token']
            try:
                file_name = os.path.basename(self.current_excel_file_path)
                date_part = os.path.splitext(file_name)[0]
                sheet_date = datetime.strptime(date_part, '%m-%d-%Y').date()
                full_datetime = datetime.combine(sheet_date, time.toPython())
                db_timestamp = full_datetime.strftime("%Y-%m-%d %H:%M:%S")
                db_event_date = full_datetime.strftime("%Y-%m-%d")
                log_tap_event(staff_name, token, db_timestamp, db_event_date)
            except Exception as e:
                QMessageBox.critical(self, "Date Error", f"Could not create timestamp for database log: {e}")
                return
            status, row_to_highlight = self.record_manual_entry(self.current_excel_file_path, staff_name, time, action)
            if "Error:" in status:
                QMessageBox.warning(self, "Action Blocked", status)
                return
            self.display_excel_content(self.current_excel_file_path)
            if row_to_highlight != -1:
                highlight_color = QColor(c.WIN_COLOR_ACCENT_PRIMARY)
                self.highlight_row(row_to_highlight, highlight_color)
            QMessageBox.information(self, "Success", f"'{staff_name}' has been manually recorded.")

    def _delete_selected_entry(self):
        row_index = self.currentRow()
        if row_index < 0: return
        item = self.item(row_index, 0)
        if not item or not item.text(): return
        name = item.text()
        reply = QMessageBox.question(self, "Confirm Deletion",
                                     f"Are you sure you want to delete the entry for '{name}'?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                wb = openpyxl.load_workbook(self.current_excel_file_path)
                ws = wb.active
                ws.protection.password = "lsst1234"
                ws.protection.sheet = False
                ws.delete_rows(row_index + 3)
                ws.protection.sheet = True
                wb.save(self.current_excel_file_path)
                self.display_excel_content(self.current_excel_file_path)
            except Exception as e:
                QMessageBox.critical(self, "Error Deleting", f"Could not delete entry from Excel: {e}")

    def _view_history_selected_entry(self):
        row_index = self.currentRow()
        if row_index < 0: return
        item = self.item(row_index, 0)
        if not item or not item.text(): return
        name = item.text()
        try:
            file_name = os.path.basename(self.current_excel_file_path)
            date_part = os.path.splitext(file_name)[0]
            query_date = datetime.strptime(date_part, '%m-%d-%Y').strftime('%Y-%m-%d')
        except Exception as e:
            QMessageBox.critical(self, "Date Error", f"Could not determine date from file name: {e}")
            return
        tap_times = get_taps_for_staff_and_date(name, query_date)
        history_dialog = StaffHistoryDialog(name, tap_times, self)
        history_dialog.exec()